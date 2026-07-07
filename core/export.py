import io
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_SECTION_LABELS = ["Pre-conditions", "Test Steps", "Expected Result"]

_PRIORITY_FILLS = {
    "critical": "DC3545",
    "high": "FD7E14",
    "normal": "6C757D",
    "low": "ADB5BD",
}
_STATUS_FILLS = {
    "ready for review": "FD7E14",
    "approved": "198754",
    "rejected": "DC3545",
    "complete": "0D6EFD",
    "superseded": "6C757D",
}

_HEADERS = [
    "#", "Prefix", "Title", "Use Case Ref", "Priority", "Status",
    "Pre-conditions", "Test Steps", "Expected Result", "Run Date",
]
_COLUMN_WIDTHS = [6, 10, 32, 16, 12, 16, 30, 45, 30, 16]


def _split_sections(text: str) -> dict:
    """Best-effort split of a test case body into Pre-conditions / Test Steps / Expected Result."""
    text = text or ""
    pattern = re.compile(
        r"\*\*(" + "|".join(re.escape(s) for s in _SECTION_LABELS) + r"):?\*\*",
        re.IGNORECASE,
    )
    matches = list(pattern.finditer(text))
    sections = {label: "" for label in _SECTION_LABELS}
    for i, m in enumerate(matches):
        label = next(l for l in _SECTION_LABELS if l.lower() == m.group(1).lower())
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        sections[label] = text[start:end].strip()
    return sections


def test_cases_to_excel(rows: list) -> bytes:
    """Render a list of test_cases DB rows into a formatted .xlsx workbook, returned as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    ws.append(_HEADERS)

    header_fill = PatternFill("solid", fgColor="16213E")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")

    for r, tc in enumerate(rows, start=2):
        body = tc.get("current_text") or tc.get("original_text") or ""
        sections = _split_sections(body)
        priority = tc.get("use_case_priority") or "Normal"
        status = tc.get("status") or ""
        date = (tc.get("run_date") or "")[:16].replace("T", " ")

        ws.append([
            tc.get("id"),
            tc.get("prefix_code") or "",
            tc.get("title") or "",
            tc.get("use_case_ref") or "",
            priority,
            status,
            sections["Pre-conditions"],
            sections["Test Steps"],
            sections["Expected Result"],
            date,
        ])
        for col in range(1, len(_HEADERS) + 1):
            ws.cell(row=r, column=col).alignment = wrap

        priority_fill = _PRIORITY_FILLS.get(priority.lower())
        if priority_fill:
            cell = ws.cell(row=r, column=5)
            cell.fill = PatternFill("solid", fgColor=priority_fill)
            cell.font = Font(color="FFFFFF", bold=True)

        status_fill = _STATUS_FILLS.get(status.lower())
        if status_fill:
            cell = ws.cell(row=r, column=6)
            cell.fill = PatternFill("solid", fgColor=status_fill)
            cell.font = Font(color="FFFFFF", bold=True)

    for i, width in enumerate(_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
